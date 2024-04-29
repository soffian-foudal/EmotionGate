import openpyxl
import numpy as np
import csv
import matplotlib.pyplot as plt


# Apri il file CSV in modalità scrittura
with open('labelAusiliare.csv', mode='w', newline='') as file:
    mediaHappy = 0
    countHappy = 0
    mediaAngry = 0
    countAngry = 0
    mediaSad = 0
    countSad = 0
    mediaNeutral = 0
    countNeutral = 0
    trueNeutral = 0
    fakeNeutral = 0

    # Crea l'oggetto writer CSV
    writer = csv.writer(file)

    # Scrivi l'intestazione delle colonne
    writer.writerow(['ID VIDEO', 'EMOZIONE', 'MEDIA', 'VARIANZA'])

    wb = openpyxl.load_workbook('responses_train_validation.xlsx')
    ws = wb.active

    for col in ws.iter_cols(min_row=ws.min_row, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        col_values = []
        for cell in col:
            value = cell.value
            if value is not None:
                col_values.append(value)

        # Rimuovi il valore della prima riga che probabilmente contiene l'intestazione della colonna
        videoId = col_values.pop(0)
        videoId = videoId[-3:]
        emotion = col_values.pop(0)
        emotion = emotion.split()[-1]
        print(videoId)
        print(emotion)
        print(col_values)

        # Calcola la media
        media = np.mean(col_values)
        varianza = np.var(col_values)

        # Aggiungi la riga al file CSV
        writer.writerow([videoId, emotion, media, varianza])

        if emotion == 'Happy':
            mediaHappy += media
            countHappy += 1
        elif emotion == 'Angry':
            mediaAngry += media
            countAngry += 1
        elif emotion == 'Sad':
            mediaSad += media
            countSad += 1
        elif emotion == 'Neutral':
            mediaNeutral += media
            countNeutral += 1

            # Calcola la soglia per distinguere "true neutral" da "fake neutral"
            # Scegli un valore soglia appropriato in base ai tuoi dati
            soglia_varianza = 0.5

            if varianza < soglia_varianza:
                trueNeutral += 1
            else:
                fakeNeutral += 1

    # Stampa il risultato
    print("Media Happy:", mediaHappy / countHappy)
    print("Media Angry:", mediaAngry / countAngry)
    print("Media Sad:", mediaSad / countSad)
    print("Media Neutral:", mediaNeutral / countNeutral)
    print("Numero True Neutral:", trueNeutral)
    print("Numero Fake Neutral:", fakeNeutral)


#uso di label di appoggio per creare label ufficiale

variance_threshold = 1.25


# Apre il file di input e di output
with open('labelAusiliare.csv', 'r') as input_file, open('label.csv', 'w', newline='') as output_file:
    reader = csv.reader(input_file)
    writer = csv.writer(output_file)

    # Legge le righe di input
    header = next(reader)
    video_index = header.index('ID VIDEO')
    emotion_index = header.index('EMOZIONE')
    mean_index = header.index('MEDIA')
    variance_index = header.index('VARIANZA')

    # Scrive l'intestazione nel file di output
    writer.writerow(['Video', 'EMOZIONE'])

    # Dizionario per memorizzare le emozioni per ogni video
    video_emotions = {}

    # Processa ogni riga del file di input
    for row in reader:
        video_id = row[video_index]
        emotion = row[emotion_index].lower()  # Converte l'emozione in minuscolo
        mean = float(row[mean_index])
        variance = float(row[variance_index])

        # Aggiunge la riga corrente al dizionario video_emotions
        if video_id not in video_emotions:
            video_emotions[video_id] = []

        video_emotions[video_id].append({'emotion': emotion, 'mean': mean, 'variance': variance})


    # Calcola l'emozione dominante per ogni video
    for video_id, emotions in video_emotions.items():
        max_mean = max(emotion['mean'] for emotion in emotions)

        max_mean_emotions = [emotion for emotion in emotions if emotion['mean'] == max_mean]
        second_mean = max(emotion['mean'] for emotion in emotions if emotion['mean'] != max_mean)
        second_max_mean_emotions = [emotion for emotion in emotions if emotion['mean'] == second_mean]
        dominant_emotion = max_mean_emotions[0]['emotion']

        # Se 'neutral' è l'unica emozione con la media più alta, controlla la varianza di 'neutral'
        if dominant_emotion == 'neutral' and len(max_mean_emotions) == 1:
            neutral_emotion = max_mean_emotions[0]
            print(video_id,max_mean_emotions[0], second_mean)
            if neutral_emotion['variance'] > variance_threshold:

                dominant_emotion = second_max_mean_emotions[0]['emotion']

        if dominant_emotion == 'neutral' and len(max_mean_emotions) == 2:
            dominant_emotion = max_mean_emotions[1]['emotion']
        # Scrive nel file di output
        writer.writerow([video_id, dominant_emotion])

# Definizione dei colori per le emozioni
colors = {'Happy': 'yellow', 'Angry': 'red', 'Sad': 'blue', 'Neutral': 'gray'}

# Lettura dei dati dal file2.csv
emotions = []
counts = []

with open("label.csv") as csv_file:
    reader = csv.reader(csv_file)
    next(reader)  # Salta la prima riga (intestazione)
    for row in reader:
        emotions.append(row[1].capitalize())

# Conteggio delle occorrenze delle emozioni
emotion_counts = {}
for emotion in emotions:
    if emotion in emotion_counts:
        emotion_counts[emotion] += 1
    else:
        emotion_counts[emotion] = 1

# Creazione del grafico
plt.bar(emotion_counts.keys(), emotion_counts.values(), color=[colors[emotion] for emotion in emotion_counts.keys()])
plt.xlabel("Emozione")
plt.ylabel("Numero")
plt.title("Distribuzione delle emozioni")

plt.show()